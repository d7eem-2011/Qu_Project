import pandas as pd
import openpyxl
from pandas import Series
from openpyxl import Workbook, load_workbook

from openpyxl.utils.dataframe import dataframe_to_rows

try:
    wb = load_workbook('الاعضاء.xlsx')

except Exception:
    wb = Workbook()
    df = pd.DataFrame([['name', 'number']])

    wb.save('الاعضاء.xlsx')



df = pd.DataFrame([['name', 'number', ]])

# Activate worksheet to write dataframe
active = wb.worksheets[0]



# Write dataframe to active worksheet
for x in dataframe_to_rows(df, index=False, header=False):
    active.append(x)
    print(x)
# Save workbook to write
wb.save('الاعضاء.xlsx')


